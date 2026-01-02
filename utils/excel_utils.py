# Utility functions for handling Excel files using openpyxl
import openpyxl

# Function to get the number of rows in a specified sheet of an Excel file
def get_row_count(file_path, sheet_name):
    workbook = openpyxl.load_workbook(file_path)
    sheet = workbook[sheet_name]
    return sheet.max_row

# Function to get the number of columns in a specified sheet of an Excel file
def get_col_count(file_path, sheet_name):
    workbook = openpyxl.load_workbook(file_path)
    sheet = workbook[sheet_name]
    return sheet.max_column

# Function to read data from a specific cell in an Excel file
def read_data(file_path, sheet_name, row_num, col_num):
    workbook = openpyxl.load_workbook(file_path)
    sheet = workbook[sheet_name]
    return sheet.cell(row_num, col_num).value

# Function to write data to a specific cell in an Excel file
def write_data(file_path, sheet_name, row_num, col_num, data):
    workbook = openpyxl.load_workbook(file_path)
    sheet = workbook[sheet_name]
    sheet.cell(row_num, col_num).value = data   # write data to the cell
    workbook.save(file_path)